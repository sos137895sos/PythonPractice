from openpyxl import load_workbook

wb = load_workbook(
    "/Users/ice/Desktop/GitHubRepositories/PythonPractice/UdemyClass/CH14/14-3Dodgers.xlsx")
result = []  # 創建一個空list，儲存資料

ws = wb.worksheets[0]  # 第一個工作表
for row in ws.iter_rows():  # iter_rows 按順序讀每一行
    # 要將所有內容存成list 可以使用 List comprehension 語法 :[cell.value for cell in row]
    result.append([cell.value for cell in row])  # 將得到的資料添加到result裡

print(result)  # 輸出工作表裡的所有數據

# 計算總全壘打數
hr_sum = 0
for hr in result[1:]:
    hr_sum += int(hr[11])

print(f"總全壘打數是： {hr_sum}")
